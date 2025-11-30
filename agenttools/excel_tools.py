from __future__ import annotations
import os
import json
from typing import Optional

from semantic_kernel.functions import kernel_function

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelTools:
    """Excel file operations for reading, writing, and manipulating spreadsheets."""

    def _check_openpyxl(self) -> str | None:
        """Return error JSON if openpyxl is not installed."""
        if not OPENPYXL_AVAILABLE:
            return json.dumps({
                "error": "openpyxl is not installed. Run: pip install openpyxl"
            })
        return None

    @kernel_function(description="Create a new Excel workbook with optional sheet names. Returns the path to the created file.")
    def create_workbook(self, path: str, sheet_names: Optional[str] = None) -> str:
        """
        Create a new Excel workbook.
        
        Args:
            path: Path where the Excel file will be saved (.xlsx)
            sheet_names: Comma-separated list of sheet names (optional). Default creates 'Sheet1'.
        """
        err = self._check_openpyxl()
        if err:
            return err
        try:
            parent_dir = os.path.dirname(path)
            if parent_dir:
                os.makedirs(parent_dir, exist_ok=True)
            wb = openpyxl.Workbook()
            
            if sheet_names:
                names = [n.strip() for n in sheet_names.split(",")]
                # Rename default sheet
                wb.active.title = names[0]
                # Add additional sheets
                for name in names[1:]:
                    wb.create_sheet(title=name)
            
            wb.save(path)
            return json.dumps({
                "status": "ok",
                "path": os.path.abspath(path),
                "sheets": [s.title for s in wb.worksheets]
            })
        except Exception as e:
            return json.dumps({"error": str(e), "path": path})

    @kernel_function(description="List all sheet names in an Excel workbook.")
    def list_sheets(self, path: str) -> str:
        """Get all sheet names from an Excel file."""
        err = self._check_openpyxl()
        if err:
            return err
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            return json.dumps({
                "status": "ok",
                "path": os.path.abspath(path),
                "sheets": sheets
            })
        except Exception as e:
            return json.dumps({"error": str(e), "path": path})

    @kernel_function(description="Read data from an Excel sheet. Returns cell data as a 2D list. Specify sheet name or uses first sheet. Optionally specify a range like 'A1:D10'.")
    def read_sheet(self, path: str, sheet_name: Optional[str] = None, cell_range: Optional[str] = None, max_rows: int = 1000) -> str:
        """
        Read data from an Excel sheet.
        
        Args:
            path: Path to the Excel file
            sheet_name: Name of the sheet to read (optional, defaults to first sheet)
            cell_range: Cell range to read like 'A1:D10' (optional, reads all if not specified)
            max_rows: Maximum number of rows to read (default 1000)
        """
        err = self._check_openpyxl()
        if err:
            return err
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    wb.close()
                    return json.dumps({
                        "error": f"Sheet '{sheet_name}' not found",
                        "available_sheets": wb.sheetnames
                    })
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            data = []
            if cell_range:
                for row in ws[cell_range]:
                    row_data = [cell.value for cell in row]
                    data.append(row_data)
            else:
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= max_rows:
                        break
                    data.append(list(row))
            
            wb.close()
            return json.dumps({
                "status": "ok",
                "path": os.path.abspath(path),
                "sheet": ws.title,
                "row_count": len(data),
                "data": data
            })
        except Exception as e:
            return json.dumps({"error": str(e), "path": path})

    @kernel_function(description="Write a value to a specific cell in an Excel sheet. Cell should be like 'A1', 'B2', etc.")
    def write_cell(self, path: str, cell: str, value: str, sheet_name: Optional[str] = None) -> str:
        """
        Write a value to a specific cell.
        
        Args:
            path: Path to the Excel file
            cell: Cell reference like 'A1', 'B2'
            value: Value to write to the cell
            sheet_name: Name of the sheet (optional, defaults to first sheet)
        """
        err = self._check_openpyxl()
        if err:
            return err
        try:
            parent_dir = os.path.dirname(path)
            if parent_dir:
                os.makedirs(parent_dir, exist_ok=True)
            
            if os.path.exists(path):
                wb = openpyxl.load_workbook(path)
            else:
                wb = openpyxl.Workbook()
            
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    wb.create_sheet(title=sheet_name)
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            # Try to convert to number if possible
            try:
                if isinstance(value, str) and "." in value:
                    value = float(value)
                elif isinstance(value, str):
                    value = int(value)
            except (ValueError, TypeError):
                pass  # Keep as string
            
            ws[cell] = value
            wb.save(path)
            wb.close()
            
            return json.dumps({
                "status": "ok",
                "path": os.path.abspath(path),
                "sheet": ws.title,
                "cell": cell,
                "value": str(value)
            })
        except Exception as e:
            return json.dumps({"error": str(e), "path": path})

    @kernel_function(description="Write multiple rows of data to an Excel sheet starting at a specific cell. Data should be a JSON array of arrays.")
    def write_rows(self, path: str, data_json: str, start_cell: str = "A1", sheet_name: Optional[str] = None) -> str:
        """
        Write multiple rows of data to a sheet.
        
        Args:
            path: Path to the Excel file
            data_json: JSON string representing a 2D array, e.g., '[["Name","Age"],["Alice",30],["Bob",25]]'
            start_cell: Starting cell for the data (default 'A1')
            sheet_name: Name of the sheet (optional)
        """
        err = self._check_openpyxl()
        if err:
            return err
        try:
            data = json.loads(data_json)
            if not isinstance(data, list):
                return json.dumps({"error": "data_json must be a JSON array of arrays"})
            
            parent_dir = os.path.dirname(path)
            if parent_dir:
                os.makedirs(parent_dir, exist_ok=True)
            
            if os.path.exists(path):
                wb = openpyxl.load_workbook(path)
            else:
                wb = openpyxl.Workbook()
            
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    wb.create_sheet(title=sheet_name)
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            # Parse start cell to get row and column
            from openpyxl.utils import coordinate_from_string, column_index_from_string
            col_str, start_row = coordinate_from_string(start_cell)
            start_col = column_index_from_string(col_str)
            
            rows_written = 0
            for row_idx, row_data in enumerate(data):
                if isinstance(row_data, list):
                    for col_idx, value in enumerate(row_data):
                        ws.cell(row=start_row + row_idx, column=start_col + col_idx, value=value)
                    rows_written += 1
            
            wb.save(path)
            wb.close()
            
            return json.dumps({
                "status": "ok",
                "path": os.path.abspath(path),
                "sheet": ws.title,
                "rows_written": rows_written,
                "start_cell": start_cell
            })
        except json.JSONDecodeError as e:
            return json.dumps({"error": f"Invalid JSON: {str(e)}"})
        except Exception as e:
            return json.dumps({"error": str(e), "path": path})

    @kernel_function(description="Add a new sheet to an existing Excel workbook.")
    def add_sheet(self, path: str, sheet_name: str) -> str:
        """Add a new sheet to an existing workbook."""
        err = self._check_openpyxl()
        if err:
            return err
        try:
            wb = openpyxl.load_workbook(path)
            if sheet_name in wb.sheetnames:
                wb.close()
                return json.dumps({
                    "error": f"Sheet '{sheet_name}' already exists",
                    "existing_sheets": wb.sheetnames
                })
            wb.create_sheet(title=sheet_name)
            wb.save(path)
            sheets = wb.sheetnames
            wb.close()
            return json.dumps({
                "status": "ok",
                "path": os.path.abspath(path),
                "new_sheet": sheet_name,
                "all_sheets": sheets
            })
        except Exception as e:
            return json.dumps({"error": str(e), "path": path})

    @kernel_function(description="Delete a sheet from an Excel workbook.")
    def delete_sheet(self, path: str, sheet_name: str) -> str:
        """Delete a sheet from a workbook."""
        err = self._check_openpyxl()
        if err:
            return err
        try:
            wb = openpyxl.load_workbook(path)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return json.dumps({
                    "error": f"Sheet '{sheet_name}' not found",
                    "available_sheets": wb.sheetnames
                })
            del wb[sheet_name]
            wb.save(path)
            sheets = wb.sheetnames
            wb.close()
            return json.dumps({
                "status": "ok",
                "path": os.path.abspath(path),
                "deleted_sheet": sheet_name,
                "remaining_sheets": sheets
            })
        except Exception as e:
            return json.dumps({"error": str(e), "path": path})

    @kernel_function(description="Get information about an Excel file: sheet names, dimensions, and row/column counts.")
    def get_workbook_info(self, path: str) -> str:
        """Get detailed information about a workbook."""
        err = self._check_openpyxl()
        if err:
            return err
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            info = {
                "status": "ok",
                "path": os.path.abspath(path),
                "sheets": []
            }
            for ws in wb.worksheets:
                sheet_info = {
                    "name": ws.title,
                    "dimensions": ws.dimensions,
                    "max_row": ws.max_row,
                    "max_column": ws.max_column
                }
                info["sheets"].append(sheet_info)
            wb.close()
            return json.dumps(info)
        except Exception as e:
            return json.dumps({"error": str(e), "path": path})

    @kernel_function(description="Read a specific cell value from an Excel sheet.")
    def read_cell(self, path: str, cell: str, sheet_name: Optional[str] = None) -> str:
        """
        Read a specific cell value.
        
        Args:
            path: Path to the Excel file
            cell: Cell reference like 'A1', 'B2'
            sheet_name: Name of the sheet (optional)
        """
        err = self._check_openpyxl()
        if err:
            return err
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    wb.close()
                    return json.dumps({
                        "error": f"Sheet '{sheet_name}' not found",
                        "available_sheets": wb.sheetnames
                    })
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            value = ws[cell].value
            wb.close()
            
            return json.dumps({
                "status": "ok",
                "path": os.path.abspath(path),
                "sheet": ws.title,
                "cell": cell,
                "value": value
            })
        except Exception as e:
            return json.dumps({"error": str(e), "path": path})

    @kernel_function(description="Clear contents of a range of cells in an Excel sheet. Range should be like 'A1:D10'.")
    def clear_range(self, path: str, cell_range: str, sheet_name: Optional[str] = None) -> str:
        """
        Clear a range of cells.
        
        Args:
            path: Path to the Excel file
            cell_range: Range to clear like 'A1:D10'
            sheet_name: Name of the sheet (optional)
        """
        err = self._check_openpyxl()
        if err:
            return err
        try:
            wb = openpyxl.load_workbook(path)
            
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    wb.close()
                    return json.dumps({
                        "error": f"Sheet '{sheet_name}' not found",
                        "available_sheets": wb.sheetnames
                    })
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            cells_cleared = 0
            for row in ws[cell_range]:
                for cell in row:
                    cell.value = None
                    cells_cleared += 1
            
            wb.save(path)
            wb.close()
            
            return json.dumps({
                "status": "ok",
                "path": os.path.abspath(path),
                "sheet": ws.title,
                "range": cell_range,
                "cells_cleared": cells_cleared
            })
        except Exception as e:
            return json.dumps({"error": str(e), "path": path})

    @kernel_function(description="Search for a value in an Excel sheet and return matching cell locations.")
    def search_value(self, path: str, search_term: str, sheet_name: Optional[str] = None, max_results: int = 100) -> str:
        """
        Search for a value in a sheet.
        
        Args:
            path: Path to the Excel file
            search_term: Value to search for (case-insensitive partial match)
            sheet_name: Name of the sheet (optional, searches active sheet)
            max_results: Maximum number of results to return (default 100)
        """
        err = self._check_openpyxl()
        if err:
            return err
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    wb.close()
                    return json.dumps({
                        "error": f"Sheet '{sheet_name}' not found",
                        "available_sheets": wb.sheetnames
                    })
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            matches = []
            search_lower = str(search_term).lower()
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        if search_lower in str(cell.value).lower():
                            matches.append({
                                "cell": cell.coordinate,
                                "value": cell.value
                            })
                            if len(matches) >= max_results:
                                break
                if len(matches) >= max_results:
                    break
            
            wb.close()
            
            return json.dumps({
                "status": "ok",
                "path": os.path.abspath(path),
                "sheet": ws.title,
                "search_term": search_term,
                "match_count": len(matches),
                "matches": matches
            })
        except Exception as e:
            return json.dumps({"error": str(e), "path": path})
